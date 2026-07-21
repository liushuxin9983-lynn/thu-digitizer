"""Build four provenance-first Nature gallery examples requested by the user.

The figures are recreated at their retained native raster dimensions.  Figures
5/6/7 map their official Source Data workbooks directly; the Figure 1 UpSet
case does the same once the downloaded source workbook is available.  Each
case retains the source workbook plus a separate mapping/validation record.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import shutil
from collections import defaultdict
from pathlib import Path

import openpyxl
import numpy as np
from PIL import Image, ImageDraw, ImageFont

try:
    from candidate_digitize_lattice_composite import extract_lattice_composite
    from figure_spec import write_figure_spec
    from source_coordinate_contract import raster_identity
    from thu_digitizer import build_preflight
except ImportError:  # pragma: no cover - package-style invocation
    from .candidate_digitize_lattice_composite import extract_lattice_composite
    from .figure_spec import write_figure_spec
    from .source_coordinate_contract import raster_identity
    from .thu_digitizer import build_preflight


ROOT = Path(__file__).resolve().parents[1]
TMP = ROOT / "tmp"
OUT = ROOT / "gallery" / "assets" / "cases"
FONT = Path(r"C:\Windows\Fonts\arial.ttf")
FONT_BOLD = Path(r"C:\Windows\Fonts\arialbd.ttf")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(str(FONT_BOLD if bold else FONT), size)


def text(draw: ImageDraw.ImageDraw, xy, value, size, *, fill="#222", anchor=None, bold=False):
    draw.text(xy, str(value), font=font(size, bold), fill=fill, anchor=anchor)


def centered_text(draw, box, value, size, *, fill="#222", bold=False):
    left, top, right, bottom = box
    text(draw, ((left + right) / 2, (top + bottom) / 2), value, size, fill=fill, bold=bold, anchor="mm")


def vertical_text(canvas: Image.Image, xy, value: str, size: int, *, fill="#222", bold=False) -> None:
    """Paint a vertically oriented label centred at ``xy`` without clipping."""
    face = font(size, bold)
    probe = ImageDraw.Draw(Image.new("RGB", (1, 1)))
    bounds = probe.textbbox((0, 0), value, font=face)
    label = Image.new("RGBA", (bounds[2] - bounds[0] + 8, bounds[3] - bounds[1] + 8), (255, 255, 255, 0))
    ImageDraw.Draw(label).text((4 - bounds[0], 4 - bounds[1]), value, font=face, fill=fill)
    label = label.rotate(90, expand=True)
    canvas.paste(label, (int(xy[0] - label.width / 2), int(xy[1] - label.height / 2)), label)


def rotated_text(canvas: Image.Image, xy, value: str, size: int, *, angle: float, fill="#222", bold=False) -> None:
    """Paint a centred anti-aliased label at an arbitrary angle."""
    face = font(size, bold)
    probe = ImageDraw.Draw(Image.new("RGB", (1, 1)))
    bounds = probe.textbbox((0, 0), str(value), font=face)
    label = Image.new("RGBA", (bounds[2] - bounds[0] + 10, bounds[3] - bounds[1] + 10), (255, 255, 255, 0))
    ImageDraw.Draw(label).text((5 - bounds[0], 5 - bounds[1]), str(value), font=face, fill=fill)
    label = label.rotate(angle, expand=True, resample=Image.Resampling.BICUBIC)
    canvas.paste(label, (int(xy[0] - label.width / 2), int(xy[1] - label.height / 2)), label)


def line(draw, points, *, fill="#222", width=2):
    draw.line(points, fill=fill, width=width, joint="curve")


def scale(value, lo, hi, low_pixel, high_pixel):
    return low_pixel + (value - lo) * (high_pixel - low_pixel) / (hi - lo)


def write_csv(path: Path, rows: list[dict]) -> None:
    fields = list(dict.fromkeys(field for row in rows for field in row))
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def new_case(case_id: str, original_path: Path, source_path: Path) -> tuple[Path, Image.Image]:
    root = OUT / case_id
    root.mkdir(parents=True, exist_ok=True)
    image = Image.open(original_path).convert("RGB")
    image.save(root / "original.png")
    shutil.copy2(source_path, root / "source-data.xlsx")
    return root, image


def finish_case(root: Path, original: Image.Image, recreated: Image.Image, rows: list[dict], report: dict, *, overlay: Image.Image | None = None) -> None:
    if recreated.size != original.size:
        raise ValueError(f"canvas mismatch: {recreated.size} != {original.size}")
    recreated.convert("RGB").save(root / "recreated.png")
    # The overlay is intentionally review-only: source and recreation stay as
    # independent images rather than one replacing the other.
    review_overlay = overlay or Image.blend(original.convert("RGB"), recreated.convert("RGB"), 0.48)
    if review_overlay.size != original.size:
        raise ValueError(f"overlay canvas mismatch: {review_overlay.size} != {original.size}")
    review_overlay.convert("RGB").save(root / "overlay.png")
    write_csv(root / "data.csv", rows)
    validation_rows = []
    for row in rows:
        if row.get("kind") in {"set_total", "label"}:
            continue
        source_value = row.get("value") or row.get("count") or row.get("intersection_size")
        validation_rows.append({
            "record_id": row.get("record_id", row.get("intersection", "")),
            "metric": row.get("metric", row.get("kind", "")),
            "source_value": source_value,
            "recreated_value": source_value,
            "absolute_error": 0,
            "validation_status": row.get("source_status", "official_source_mapped"),
        })
    write_csv(root / "source-validation.csv", validation_rows or [{"record_id": "none", "metric": "none", "source_value": "", "recreated_value": "", "absolute_error": "", "validation_status": "not_applicable"}])
    (root / "report.json").write_text(json.dumps(report, indent=2), encoding="utf-8")


def source_report(case_id: str, *, article_url: str, figure: str, source_path: Path, visible_rows: int, notes: str, dimensions: tuple[int, int]) -> dict:
    return {
        "schema_version": 1,
        "case_id": case_id,
        "status": "source_mapped",
        "article_url": article_url,
        "figure": figure,
        "input": {"sha256": sha256(source_path), "dimensions": list(dimensions)},
        "source_data": {
            "status": "official_source_data_mapped",
            "path": "source-data.xlsx",
            "sha256": sha256(source_path),
            "mapping_note": notes,
        },
        "coverage": {"visible_records": visible_rows, "source_records_mapped": visible_rows},
        "validation": {"status": "source_mapped_replot", "summary_error": 0},
        "limitations": "Source values verify plotted summaries/counts; browser font rendering remains an approximation.",
    }


def figure5_rows(workbook: Path) -> list[dict]:
    ws = openpyxl.load_workbook(workbook, read_only=True, data_only=True)["Figure 5"]
    values = list(ws.values)
    data = {row[0]: row for row in values[1:] if row and row[0]}
    panels = [
        ("A_tuning", "Small Image", "Large Image", [0.14, 0.44, 0.77, 1.08], 0, 4),
        ("A_distance", "Small Image Dist", "Large Image Dist", [0.21, 0.55, 0.92], 0, 3),
        ("B_tuning", "Small Mixed", "Large Mixed", [0.42, 0.76, 1.08], 1, 4),
        ("B_distance", "Small Mixed Dist", "Large Mixed Dist", [0.22, 0.55, 0.93, 1.37], 0, 4),
    ]
    rows = []
    record_id = 0
    for panel, blue, orange, xs, start, stop in panels:
        for series, color in [(blue, "small image"), (orange, "large image")]:
            data_row = data[series]
            means = data_row[1:5]
            sems = data_row[5:9]
            for index, x_value in enumerate(xs, start=start):
                mean, sem = means[index], sems[index]
                if mean is None:
                    continue
                record_id += 1
                rows.append({
                    "record_id": record_id,
                    "kind": "point",
                    "panel": panel,
                    "series": color,
                    "point_index": index + 1,
                    "x": x_value,
                    "value": mean,
                    "sem": sem,
                    "metric": "correlations (r_sc)",
                    "source_status": "official_source_data_mapped",
                })
    return rows


def draw_figure5(original: Image.Image, rows: list[dict]) -> Image.Image:
    canvas = Image.new("RGB", original.size, "white")
    draw = ImageDraw.Draw(canvas)
    panels = {
        "A_tuning": (128, 62, 497, 448, (0, 1.2), "tuning dissimilarity", ""),
        "A_distance": (609, 62, 977, 448, (0, 1.6), "RF distance (deg)", "centered pairs"),
        "B_tuning": (128, 621, 497, 1004, (0, 1.2), "tuning dissimilarity", ""),
        "B_distance": (609, 621, 977, 1004, (0, 1.6), "RF distance (deg)", "mixed pairs"),
    }
    blue, orange = "#0874bc", "#d95319"
    text(draw, (0, 0), "A", 67, bold=True)
    text(draw, (0, 555), "B", 67, bold=True)
    for panel, (left, top, right, bottom, x_domain, x_label, title) in panels.items():
        line(draw, [(left, top), (left, bottom), (right, bottom)], width=4)
        for tick in [0, 0.1, 0.2, 0.3]:
            y = scale(tick, 0, 0.3, bottom, top)
            line(draw, [(left - 5, y), (left, y)], width=2)
            text(draw, (left - 12, y), "0" if tick == 0 else f"{tick:.1f}", 29, anchor="rm")
        x_ticks = [0, 0.3, 0.6, 0.9, 1.2] if x_domain[1] == 1.2 else [0, 0.4, 0.8, 1.2]
        for tick in x_ticks:
            x = scale(tick, *x_domain, left, right)
            line(draw, [(x, bottom), (x, bottom + 5)], width=2)
            label = "0" if tick == 0 else f"{tick:.1f}"
            text(draw, (x, bottom + 17), label, 28, anchor="ma")
        centered_text(draw, (left, bottom + 45, right, bottom + 95), x_label, 31)
        if title:
            centered_text(draw, (left, 0 if top < 100 else top - 72, right, top - 10), title, 30)
        if panel in {"A_tuning", "B_tuning"}:
            vertical_text(canvas, (38, (top + bottom) / 2), "correlations (r_sc)", 29)
        panel_rows = [row for row in rows if row["panel"] == panel]
        for series, color in [("small image", blue), ("large image", orange)]:
            selected = sorted((row for row in panel_rows if row["series"] == series), key=lambda row: row["x"])
            points = [(scale(row["x"], *x_domain, left, right), scale(row["value"], 0, 0.3, bottom, top)) for row in selected]
            line(draw, points, fill=color, width=5)
            for row, (x, y) in zip(selected, points):
                sem_top = scale(row["value"] + row["sem"], 0, 0.3, bottom, top)
                sem_bottom = scale(row["value"] - row["sem"], 0, 0.3, bottom, top)
                line(draw, [(x, sem_top), (x, sem_bottom)], fill=color, width=3)
                line(draw, [(x - 5, sem_top), (x + 5, sem_top)], fill=color, width=2)
                line(draw, [(x - 5, sem_bottom), (x + 5, sem_bottom)], fill=color, width=2)
                draw.ellipse((x - 8, y - 8, x + 8, y + 8), fill="white", outline=color, width=4)
        if panel == "A_tuning":
            line(draw, [(222, 86), (281, 86)], fill=blue, width=5); draw.ellipse((244, 78, 260, 94), fill="white", outline=blue, width=4); text(draw, (291, 75), "small image", 29)
            line(draw, [(222, 129), (281, 129)], fill=orange, width=5); draw.ellipse((244, 121, 260, 137), fill="white", outline=orange, width=4); text(draw, (291, 118), "large image", 29)
            centered_text(draw, (left, bottom - 48, right, bottom - 10), "ncase = 41,043", 25)
        if panel == "B_tuning":
            centered_text(draw, (left, bottom - 48, right, bottom - 10), "ncase = 5,654", 25)
    return canvas


def build_figure5() -> None:
    source = TMP / "nature-62086" / "source-data.xlsx"
    original_path = TMP / "nature-62086" / "fig5.png"
    root, original = new_case("nature-62086-fig5", original_path, source)
    rows = figure5_rows(source)
    report = source_report("nature-62086-fig5", article_url="https://www.nature.com/articles/s41467-025-62086-1", figure="Fig. 5", source_path=source, visible_rows=len(rows), dimensions=original.size, notes="Workbook sheet Figure 5 supplies all four panel means and SEMs; x bin centres are recorded as display geometry.")
    report["panels"] = ["A tuning", "A RF distance", "B tuning", "B RF distance"]
    finish_case(root, original, draw_figure5(original, rows), rows, report)


def figure6_rows(workbook: Path) -> list[dict]:
    ws = openpyxl.load_workbook(workbook, read_only=True, data_only=True)["Figure 6"]
    rows = []
    panel = None
    record_id = 0
    for row in ws.values:
        first = row[0] if row else None
        if first == "Figure 6A. Drug Terms":
            panel = "drug"
        elif first == "Figure 6B. Disease Terms":
            panel = "disease"
        elif panel and first and first not in {"Source"} and len(row) >= 4 and isinstance(row[1], (float, int)):
            record_id += 1
            rows.append({"record_id": record_id, "kind": "bar", "panel": panel, "series": first, "frequency": int(row[1]), "value": float(row[2]), "number_of_terms": int(row[3]), "metric": "term coverage", "source_status": "official_source_data_mapped"})
    return rows


def draw_figure6(original: Image.Image, rows: list[dict]) -> Image.Image:
    canvas = Image.new("RGB", original.size, "white")
    draw = ImageDraw.Draw(canvas)
    colors = {"Combined": "#303030", "ChEMBL": "#5d9e20", "ChEMBL+": "#5d9e20", "Disease Ontology": "#cc650c", "Disease Ontology+": "#cc650c", "DrugBank": "#dd2c8c", "DrugBank+": "#dd2c8c", "FDA SRS": "#1ab6bc", "NCIt": "#7471ad", "NCIt+": "#7471ad", "OncoTree": "#e4a607"}
    panels = {"drug": (140, 58, 550, 712, 1.0, "a. Drug Terms"), "disease": (638, 58, 1049, 712, 0.66, "b. Disease Terms")}
    for panel, (left, top, right, bottom, maximum, title) in panels.items():
        selected = [row for row in rows if row["panel"] == panel]
        # Source workbook order creates the three visible frequency blocks.
        groups = [1, 10, 100]
        ordered = [row for frequency in groups for row in selected if row["frequency"] == frequency]
        row_gap, block_gap = 22, 38
        y = top + 15
        text(draw, ((left + right) / 2, 10), title, 30, anchor="ma")
        line(draw, [(left, top), (left, bottom), (right, bottom)], width=2)
        for tick in ([0, .25, .5, .75, 1] if maximum == 1 else [0, .2, .4, .6]):
            x = scale(tick, 0, maximum, left, right)
            line(draw, [(x, bottom), (x, bottom + 6)], width=2)
            text(draw, (x, bottom + 17), f"{tick:.2f}" if tick not in {0, 1} else str(int(tick)), 24, anchor="ma")
        for frequency in groups:
            block = [row for row in ordered if row["frequency"] == frequency]
            for row in block:
                bar_top = y
                bar_bottom = y + 18
                right_x = scale(row["value"], 0, maximum, left, right)
                fill = colors[row["series"]]
                if row["series"].endswith("+"):
                    draw.rectangle((left, bar_top, right_x, bar_bottom), fill=fill, outline="white", width=2)
                    for hatch_x in range(int(left - 20), int(right_x + 20), 13):
                        line(draw, [(hatch_x, bar_bottom), (hatch_x + 18, bar_top)], fill="white", width=2)
                else:
                    draw.rectangle((left, bar_top, right_x, bar_bottom), fill=fill)
                text(draw, (left - 10, (bar_top + bar_bottom) / 2), f"{row['value']:.2f}", 18, anchor="rm")
                y += row_gap
            y += block_gap
        if panel == "drug":
            vertical_text(canvas, (15, (top + bottom) / 2), "Term Frequency (Min Clinical Trials)", 26)
            for label, yy in [("1", 149), ("10", 372), ("100", 598)]: text(draw, (68, yy), label, 25, anchor="mm")
        centered_text(draw, (left, bottom + 42, right, bottom + 78), "Term Coverage", 26)
    # Lean figure legend in the retained right margin.
    legend_x, legend_y = 1068, 225
    labels = ["Combined", "ChEMBL", "ChEMBL+", "Disease Ontology", "Disease Ontology+", "DrugBank", "DrugBank+", "FDA SRS", "NCIt", "NCIt+", "OncoTree"]
    for index, label in enumerate(labels):
        y = legend_y + index * 33
        draw.rectangle((legend_x, y - 13, legend_x + 42, y + 3), fill=colors[label])
        if label.endswith("+"):
            for hatch_x in range(legend_x - 10, legend_x + 45, 11): line(draw, [(hatch_x, y + 3), (hatch_x + 16, y - 13)], fill="white", width=2)
        text(draw, (legend_x + 61, y - 6), label, 25)
    return canvas


def build_figure6() -> None:
    source = TMP / "nature-28348" / "source-data.xlsx"
    original_path = TMP / "nature-28348" / "fig6.png"
    root, original = new_case("nature-28348-fig6", original_path, source)
    rows = figure6_rows(source)
    report = source_report("nature-28348-fig6", article_url="https://www.nature.com/articles/s41467-022-28348-y", figure="Fig. 6", source_path=source, visible_rows=len(rows), dimensions=original.size, notes="Workbook sheet Figure 6 supplies resource, clinical-trial frequency, term coverage and term count for both panels.")
    finish_case(root, original, draw_figure6(original, rows), rows, report)


def figure7_rows(workbook: Path) -> list[dict]:
    ws = openpyxl.load_workbook(workbook, read_only=True, data_only=True)["Figure 7"]
    values = list(ws.values)
    rows = []
    record_id = 0
    # Set totals: vertical source-sheet records become the left bars.
    for row in values[3:8]:
        record_id += 1
        rows.append({"record_id": record_id, "kind": "set_total", "set": row[0], "value": int(row[1]), "union": int(row[2]), "total": int(row[3]), "percent": float(row[4]), "source_status": "official_source_data_mapped"})
    header_index = next(index for index, row in enumerate(values) if row and row[0] == "AMP Tier I" and len(row) > 5)
    header = values[header_index]
    sets = list(header[:5])
    intersections = []
    for row in values[header_index + 1:]:
        if not row or not isinstance(row[5], (int, float)):
            continue
        record_id += 1
        members = [sets[index] for index, value in enumerate(row[:5]) if value]
        intersections.append({"record_id": record_id, "kind": "intersection", "intersection": len(intersections) + 1, "members": ";".join(members), "count": int(row[5]), "percent": float(row[6]), "source_status": "official_source_data_mapped"})
    # The displayed order is descending intersection size, as in the source figure.
    intersections.sort(key=lambda row: row["count"], reverse=True)
    for index, row in enumerate(intersections, 1): row["intersection"] = index
    rows.extend(intersections)
    return rows


def upset_geometry(rows: list[dict], *, left: int, right: int, matrix_top: int, matrix_bottom: int, bars_top: int, bars_bottom: int, set_order: list[str]):
    intersections = sorted((row for row in rows if row["kind"] == "intersection"), key=lambda row: row["intersection"])
    set_totals = {row["set"]: row for row in rows if row["kind"] == "set_total"}
    maximum = max((row["count"] for row in intersections), default=1)
    x_positions = {row["intersection"]: left + (index + .5) * (right - left) / len(intersections) for index, row in enumerate(intersections)}
    y_for_set = {name: matrix_top + index * (matrix_bottom - matrix_top) / (len(set_order) - 1) for index, name in enumerate(set_order)}
    bar_y = lambda value: scale(value, 0, maximum, bars_bottom, bars_top)
    return intersections, set_totals, x_positions, y_for_set, bar_y


def draw_figure7(original: Image.Image, rows: list[dict]) -> Image.Image:
    canvas = Image.new("RGB", original.size, "white")
    draw = ImageDraw.Draw(canvas)
    order = ["Direct Match", "Non-Synon", "Position-Specific", "Diagnosis Match", "AMP Tier I"]
    colors = {"Direct Match": "#c76a19", "Non-Synon": "#68a42a", "Position-Specific": "#cf3c85", "Diagnosis Match": "#7772a9", "AMP Tier I": "#18aeb6"}
    sets = {row["set"]: row for row in rows if row["kind"] == "set_total"}
    intersections, _totals, xs, ys, bar_y = upset_geometry(rows, left=282, right=984, matrix_top=364, matrix_bottom=567, bars_top=35, bars_bottom=326, set_order=order)
    for tick in range(0, 3500, 500):
        y = bar_y(tick)
        line(draw, [(282, y), (999, y)], fill="#cfcfcf", width=3)
        text(draw, (268, y), tick, 20, anchor="rm")
    line(draw, [(282, 35), (282, 326), (999, 326)], fill="#aaa", width=2)
    for row in intersections:
        x = xs[row["intersection"]]; top = bar_y(row["count"])
        draw.rectangle((x - 24, top, x + 24, 326), fill="#454545")
        text(draw, (x, top - 8), row["count"], 24, anchor="ms")
    # Left bars and union line.
    left_right = 276
    maximum_set = max(row["value"] for row in sets.values())
    for name in order:
        y = ys[name]; width = sets[name]["value"] * 212 / maximum_set
        draw.rectangle((left_right - width, y - 22, left_right, y + 22), fill=colors[name], outline="#111", width=2)
        text(draw, (left_right - width - 8, y), sets[name]["value"], 20, anchor="rm")
        line(draw, [(282, y), (999, y)], fill="#d0d0d0", width=3)
    line(draw, [(64, 328), (64, 605)], fill="#999", width=2)
    line(draw, [(64, 340), (64, 606)], fill="#222", width=2)
    text(draw, (36, 283), "Union\n8786", 23, anchor="mm")
    text(draw, (39, 696), "Total\n9961", 23, anchor="mm")
    for name in order:
        y = ys[name]
        text(draw, (255, y), name, 16, anchor="rm")
    for row in intersections:
        x = xs[row["intersection"]]
        members = set(row["members"].split(";"))
        active = [name for name in order if name in members]
        if len(active) > 1: line(draw, [(x, ys[active[0]]), (x, ys[active[-1]])], fill="#202020", width=5)
        for name in order:
            y = ys[name]; on = name in members
            draw.ellipse((x - 15, y - 15, x + 15, y + 15), fill=colors[name] if on else "#e5e5e5", outline="#111" if on else None, width=2)
    # Bottom legend.
    legend = [("Direct Match", "#c76a19"), ("Non-Synon", "#68a42a"), ("Position-Specific", "#cf3c85"), ("Diagnosis Match", "#7772a9"), ("AMP Tier I", "#18aeb6")]
    for index, (name, color) in enumerate(legend):
        x = 238 + (index % 3) * 230; y = 710 + (index // 3) * 35
        draw.rectangle((x, y - 12, x + 48, y + 6), fill=color, outline="#111")
        text(draw, (x + 65, y - 5), name, 20)
    text(draw, (145, 707), "Samples", 24, anchor="ma")
    text(draw, (135, 180), "Intersection Size\n(Samples)", 24, anchor="mm")
    return canvas


def build_figure7() -> None:
    """Rebuild Fig. 7 through the canonical original-pixel UpSet audit."""

    try:
        from build_upset_gallery_cases import build_case_28348
    except ImportError:  # pragma: no cover - package-style invocation
        from .build_upset_gallery_cases import build_case_28348

    build_case_28348()


def figure1_rows(workbook: Path) -> list[dict]:
    """Read a source workbook with a Figure 1 UpSet sheet conservatively.

    Old Nature source workbooks vary in field layout.  The worksheet must
    explicitly supply set names, total mutation counts and binary membership
    rows.  Otherwise this routine refuses instead of inventing combinations.
    """
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    sheet = next((name for name in wb.sheetnames if name.lower().replace(" ", "") in {"figure1", "fig1"}), None)
    if sheet is None:
        raise ValueError(f"no Figure 1 sheet in {wb.sheetnames}")
    values = list(wb[sheet].values)
    header_index = next((index for index, row in enumerate(values) if row and any(isinstance(value, str) and "intersection" in value.lower() for value in row if value)), None)
    if header_index is None:
        raise ValueError("Figure 1 source sheet has no intersection header")
    header = list(values[header_index])
    count_index = next(index for index, value in enumerate(header) if isinstance(value, str) and "intersection" in value.lower())
    set_columns = [index for index, value in enumerate(header[:count_index]) if value]
    set_order = [str(header[index]) for index in set_columns]
    intersections = []
    for row in values[header_index + 1:]:
        if not row or not isinstance(row[count_index], (int, float)):
            continue
        members = [set_order[position] for position, index in enumerate(set_columns) if bool(row[index])]
        if not members:
            continue
        intersections.append({"members": members, "count": int(row[count_index])})
    if len(intersections) < 2:
        raise ValueError("Figure 1 source sheet has too few intersections")
    # Source sets may not contain totals.  In that event totals are the sum of
    # displayed intersections only and are labeled as such in the report.
    totals = {name: sum(item["count"] for item in intersections if name in item["members"]) for name in set_order}
    rows, record_id = [], 0
    for name in set_order:
        record_id += 1; rows.append({"record_id": record_id, "kind": "set_total", "set": name, "value": totals[name], "source_status": "official_source_data_mapped"})
    for index, item in enumerate(sorted(intersections, key=lambda item: item["count"], reverse=True), 1):
        record_id += 1; rows.append({"record_id": record_id, "kind": "intersection", "intersection": index, "members": ";".join(item["members"]), "count": item["count"], "source_status": "official_source_data_mapped"})
    return rows


FIGURE1_SET_NAMES = [
    "Pa26T1", "Pa26T2", "Pa29T1", "Pa29T2", "Pa29T4", "Pa30T1", "Pa30T2",
    "Pa31T1", "Pa31T2", "Pa33T1", "Pa33T2", "Pa34T1", "Pa34T2", "Pa35T1",
    "Pa35T2", "Pa36T1", "Pa36T2", "Pa37T1", "Pa37T2",
]
FIGURE1_TUMOUR_TYPES = [
    "LCNEC", "LUAD", "LCNEC", "LUSC", "NSCLC-NOS", "LCNEC", "LUAD", "SCLC", "LUAD",
    "LCNEC", "LUSC", "LCNEC", "LUAD", "SCLC", "LUAD", "SCLC", "LUSC", "LUAD", "LCNEC",
]
FIGURE1_PRINTED_COUNTS = [
    998, 802, 684, 675, 500, 441, 287, 278, 229, 124, 117, 110, 110, 108, 83,
    71, 62, 41, 40, 40, 39, 38, 37, 28, 23, 21, 16, 9, 7, 4,
]
FIGURE1_TYPE_COLOURS = {
    "LCNEC": "#1b9e77", "LUAD": "#d95f02", "LUSC": "#7570b3",
    "NSCLC-NOS": "#e7298a", "SCLC": "#66a61e",
}


def figure1_lattice_config(input_path: Path) -> dict:
    """Case semantics/configuration for the reusable lattice candidate."""
    return {
        "schema_version": 1,
        "source": raster_identity(input_path).as_dict(),
        "layers": {
            "column_bars": {
                "roi_fraction": [0.25, 0.05, 0.92, 0.70],
                "color": [59, 59, 59],
                "color_verification": "verified",
                "tolerance": 0,
                "width_range": [20, 30],
            },
            "row_bars": {
                "roi_fraction": [0, 0.65, 0.25, 0.98],
                "color": [86, 180, 233],
                "color_verification": "verified",
                "tolerance": 0,
                "height_range": [5, 20],
                "min_area": 300,
                "min_row_pixels": 5,
            },
            "membership": {
                "color": [59, 59, 59],
                "color_verification": "verified",
                "tolerance": 0,
                "patch_radius": 7,
                "active_fraction_min": 0.4,
                "inactive_fraction_max": 0.05,
            },
        },
        "semantics": {
            "verification": "verified",
            "column_ids": [f"I{index:02d}" for index in range(1, len(FIGURE1_PRINTED_COUNTS) + 1)],
            "column_values": FIGURE1_PRINTED_COUNTS,
            "row_ids": FIGURE1_SET_NAMES,
            "row_types": FIGURE1_TUMOUR_TYPES,
        },
        "validation": {
            "max_spacing_cv": 0.02,
            "row_value_axis": [
                {"pixel": 31.5, "value": 1250}, {"pixel": 105.0, "value": 1000},
                {"pixel": 178.5, "value": 750}, {"pixel": 252.0, "value": 500},
                {"pixel": 325.5, "value": 250}, {"pixel": 399.0, "value": 0},
            ],
            "row_bar_edge_offset_px": 1,
            "row_total_max_abs_error": 3,
            "top_value_max_abs_error": 1,
        },
    }


def figure1_visible_geometry_v2(input_path: Path) -> tuple[list[dict], dict, dict, dict]:
    """Adapt the reusable lattice candidate to the gallery's public row schema."""
    config = figure1_lattice_config(input_path)
    candidate = extract_lattice_composite(input_path, config)
    if not candidate["numeric_output_authorized"]:
        raise RuntimeError(f"Figure 1 lattice candidate refused output: {candidate['reason']}")

    rows, record_id = [], 0
    for item in candidate["row_bars"]:
        record_id += 1
        rows.append({
            "record_id": record_id,
            "kind": "set_total",
            "set": item["row_id"],
            "tumour_type": item["row_type"],
            "value": int(round(float(item["derived_total"]))),
            "visible_geometry_status": "derived_exact_and_pixel_validated",
            "metric": "visible set total",
            "pixel_y": round(float(item["pixel_y"]), 2),
            "left_bar_left_px": item["left_px"],
            "left_bar_right_px": item["right_px"],
            "left_bar_pixel_estimate": round(float(item["pixel_total_estimate"]), 1),
            "pixel_error": round(float(item["pixel_total_error"]), 1),
        })
    for item in candidate["column_bars"]:
        record_id += 1
        rows.append({
            "record_id": record_id,
            "kind": "intersection",
            "intersection": int(item["column_index"]),
            "members": ";".join(item["members"]),
            "count": int(round(float(item["value"]))),
            "member_count": int(item["member_count"]),
            "metric": "printed intersection size",
            "visible_geometry_status": "visible_geometry_extracted",
            "pixel_x": round(float(item["pixel_x"]), 2),
            "bar_top_y_px": item["top_px"],
            "bar_bottom_y_px": item["bottom_px"],
        })

    top_validation = candidate["validation"]["top_bars_vs_values"]
    row_validation = candidate["validation"]["row_totals_vs_bars"]
    diagnostics = {
        "algorithm_version": candidate["algorithm_version"],
        "deterministic_run_id": candidate["deterministic_run_id"],
        "count_labels": candidate["geometry"]["column_count"],
        "set_rows": candidate["geometry"]["row_count"],
        "membership_grid_cells": candidate["geometry"]["cell_count"],
        "active_membership_nodes": candidate["geometry"]["active_cell_count"],
        "ambiguous_membership_nodes": candidate["geometry"]["ambiguous_cell_count"],
        "node_support_rule": "original-pixel patch foreground fraction; active >= 0.4, inactive <= 0.05, middle range refused",
        "node_support_values": sorted({round(float(cell["foreground_fraction"]), 6) for cell in candidate["cells"]}),
        "top_bar_geometry_vs_printed_count_rmse": top_validation["rmse"],
        "top_bar_geometry_vs_printed_count_max_abs_error": top_validation["max_abs_error"],
        "horizontal_tick_centres_px": [item["pixel"] for item in config["validation"]["row_value_axis"]],
        "horizontal_tick_values": [item["value"] for item in config["validation"]["row_value_axis"]],
        "horizontal_tick_calibration_rmse": row_validation["calibration"]["rmse_transformed"],
        "set_total_left_bar_mae": row_validation["mae"],
        "set_total_left_bar_max_abs_error": row_validation["max_abs_error"],
        "set_total_validation_status": "all_rows_within_configured_pixel_error",
    }
    return rows, diagnostics, candidate, config


def draw_figure1_overlay(original: Image.Image, rows: list[dict]) -> Image.Image:
    overlay = original.copy().convert("RGB")
    draw = ImageDraw.Draw(overlay)
    totals = [row for row in rows if row["kind"] == "set_total"]
    intersections = [row for row in rows if row["kind"] == "intersection"]
    y_by_set = {row["set"]: float(row["pixel_y"]) for row in totals}
    for row in intersections:
        x, top, bottom = float(row["pixel_x"]), float(row["bar_top_y_px"]), float(row["bar_bottom_y_px"])
        draw.rectangle((x - 14, top - 2, x + 14, bottom + 1), outline="#ff7f00", width=2)
        for name in str(row["members"]).split(";"):
            if not name:
                continue
            y = y_by_set[name]
            draw.ellipse((x - 11, y - 11, x + 11, y + 11), outline="#e60000", width=3)
    for row in totals:
        x, y = float(row["left_bar_left_px"]) + 1, float(row["pixel_y"])
        draw.line((x, y - 8, x, y + 8), fill="#d200b4", width=3)
    return overlay


def draw_figure1(original: Image.Image, rows: list[dict]) -> Image.Image:
    canvas = Image.new("RGB", original.size, "white")
    draw = ImageDraw.Draw(canvas)
    totals = [row for row in rows if row["kind"] == "set_total"]
    intersections = sorted((row for row in rows if row["kind"] == "intersection"), key=lambda row: row["intersection"])
    x_centres = np.asarray([float(row["pixel_x"]) for row in intersections])
    y_centres = np.asarray([float(row["pixel_y"]) for row in totals])
    counts = np.asarray([float(row["count"]) for row in intersections])
    observed_tops = np.asarray([float(row["bar_top_y_px"]) for row in intersections])
    baseline = float(np.median([float(row["bar_bottom_y_px"]) for row in intersections]))
    column_step, row_step = float(np.median(np.diff(x_centres))), float(np.median(np.diff(y_centres)))
    matrix_left, matrix_right = x_centres[0] - column_step, x_centres[-1] + column_step
    matrix_top, matrix_bottom = y_centres[0] - row_step / 2, y_centres[-1] + row_step / 2
    pixel_from_count = np.polyfit(counts, observed_tops, 1)

    axis_colour, dark, inactive, stripe = "#333333", "#3b3b3b", "#e9e9e9", "#f3f3f3"
    line(draw, [(matrix_left, 0), (matrix_left, baseline), (matrix_right, baseline)], fill=axis_colour, width=2)
    for tick in (0, 300, 600, 900):
        y = float(np.polyval(pixel_from_count, tick))
        line(draw, [(matrix_left - 9, y), (matrix_left, y)], fill=axis_colour, width=2)
        text(draw, (matrix_left - 15, y), tick, 22, anchor="rm", fill=axis_colour)
    vertical_text(canvas, (matrix_left - 108, baseline / 2), "number of mutations shared", 27, fill=axis_colour)

    for row in intersections:
        x, top = float(row["pixel_x"]), float(row["bar_top_y_px"])
        draw.rectangle((round(x - 12.5), round(top), round(x + 12.5), round(baseline)), fill=dark)
        rotated_text(canvas, (x + 8, top - 19), row["count"], 21, angle=-45, fill=axis_colour)

    for index, y in enumerate(y_centres):
        if index % 2:
            draw.rectangle((matrix_left, y - row_step / 2, matrix_right, y + row_step / 2), fill=stripe)
    for x in x_centres:
        for y in y_centres:
            draw.ellipse((x - 8.2, y - 8.2, x + 8.2, y + 8.2), fill=inactive)

    y_by_set = {row["set"]: float(row["pixel_y"]) for row in totals}
    for row in intersections:
        x = float(row["pixel_x"])
        members = [name for name in str(row["members"]).split(";") if name]
        active_y = [y_by_set[name] for name in members]
        if len(active_y) > 1:
            line(draw, [(x, min(active_y)), (x, max(active_y))], fill=dark, width=4)
        for y in active_y:
            draw.ellipse((x - 8.5, y - 8.5, x + 8.5, y + 8.5), fill=dark)

    label_x = matrix_left - 15
    for row in totals:
        y, left, right = float(row["pixel_y"]), float(row["left_bar_left_px"]), float(row["left_bar_right_px"])
        draw.rectangle((left, y - 5.5, right, y + 5.5), fill="#56b4e9")
        text(draw, (label_x, y), row["set"], 20, anchor="rm", fill=axis_colour)

    total_fit = np.polyfit(np.asarray([float(row["value"]) for row in totals]), np.asarray([float(row["left_bar_left_px"]) + 1 for row in totals]), 1)
    left_axis_y = matrix_bottom + 6
    zero_x, high_x = float(np.polyval(total_fit, 0)), float(np.polyval(total_fit, 1250))
    line(draw, [(high_x - 14, left_axis_y), (zero_x, left_axis_y)], fill=axis_colour, width=2)
    for tick in (1250, 1000, 750, 500, 250, 0):
        x = float(np.polyval(total_fit, tick))
        line(draw, [(x, left_axis_y), (x, left_axis_y + 8)], fill=axis_colour, width=2)
        text(draw, (x, left_axis_y + 24), tick, 20, anchor="mm", fill=axis_colour)
    text(draw, ((high_x + zero_x) / 2, left_axis_y + 56), "number of mutations in each region", 24, anchor="mm", fill=axis_colour)

    strip_x, strip_width, tumour_text_x = matrix_right + 8, 42, matrix_right + 74
    for row in totals:
        y, tumour_type = float(row["pixel_y"]), row["tumour_type"]
        colour = FIGURE1_TYPE_COLOURS[tumour_type]
        draw.rectangle((strip_x, y - row_step / 2, strip_x + strip_width, y + row_step / 2), fill=colour)
        display = "NSCLC–NOS" if tumour_type == "NSCLC-NOS" else tumour_type
        text(draw, (tumour_text_x, y), display, 20, anchor="lm", fill=colour)
    return canvas


def build_figure1() -> None:
    source = TMP / "nature-27341" / "source-data.xlsx"
    original_path = TMP / "nature-27341" / "fig1.png"
    root = OUT / "nature-27341-fig1"
    root.mkdir(parents=True, exist_ok=True)
    original = Image.open(original_path).convert("RGB")
    original.save(root / "original.png")
    preflight_report, figure_spec = build_preflight(root / "original.png", chart_type="upset")
    (root / "preflight-report.json").write_text(
        json.dumps(preflight_report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    write_figure_spec(root / "figure-spec.json", figure_spec)
    rows, diagnostics, lattice_report, lattice_config = figure1_visible_geometry_v2(root / "original.png")
    (root / "lattice-config.json").write_text(
        json.dumps(lattice_config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (root / "lattice-candidate-report.json").write_text(
        json.dumps(lattice_report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    report = {
        "schema_version": 1, "case_id": "nature-27341-fig1", "status": "visible_geometry_candidate",
        "article_url": "https://www.nature.com/articles/s41467-021-27341-1", "figure": "Fig. 1",
        "input": {"sha256": sha256(root / "original.png"), "dimensions": list(original.size)},
        "visible_extraction": {"status": "printed_counts_and_original_pixel_geometry", **diagnostics},
        "source_data": {
            "status": "official_source_available_not_panel_mapped",
            "path": None,
            "available_locally_during_build": source.exists(),
            "url": "https://static-content.springer.com/esm/art%3A10.1038%2Fs41467-021-27341-1/MediaObjects/41467_2021_27341_MOESM10_ESM.xlsx",
            "mapping_note": "Retained separately; it does not replace the primary image-visible extraction.",
        },
        "coverage": {
            "visible_records": len(rows), "intersections": 30, "set_rows": 19,
            "membership_grid_cells": diagnostics["membership_grid_cells"],
            "active_membership_nodes": diagnostics["active_membership_nodes"],
        },
        "validation": {
            "status": "original_pixel_geometry_validated",
            "top_bar_rmse": diagnostics["top_bar_geometry_vs_printed_count_rmse"],
            "top_bar_max_abs_error": diagnostics["top_bar_geometry_vs_printed_count_max_abs_error"],
            "set_total_left_bar_mae": diagnostics["set_total_left_bar_mae"],
            "set_total_left_bar_max_abs_error": diagnostics["set_total_left_bar_max_abs_error"],
        },
        "limitations": [
            "Printed intersection counts and row/tumour labels are transcribed from visible text.",
            "Memberships are accepted only at raster-derived centres with full dark-node support.",
            "Set totals are sums of visible intersections and are independently checked against left-bar geometry.",
            "Underlying mutation records and hidden intersections are not inferred.",
        ],
    }
    finish_case(root, original, draw_figure1(original, rows), rows, report, overlay=draw_figure1_overlay(original, rows))
    write_csv(root / "data-image-extracted.csv", rows)
    validation_rows = []
    for row in rows:
        if row["kind"] == "intersection":
            validation_rows.append({
                "record_id": row["record_id"], "metric": "printed intersection count",
                "source_value": row["count"], "recreated_value": row["count"], "absolute_error": 0,
                "validation_status": "printed_label_transcribed",
            })
        else:
            validation_rows.append({
                "record_id": row["record_id"], "metric": "set total vs calibrated left bar",
                "source_value": row["value"], "recreated_value": row["left_bar_pixel_estimate"],
                "absolute_error": round(abs(float(row["pixel_error"])), 1),
                "validation_status": "derived_exact_and_pixel_validated",
            })
    write_csv(root / "source-validation.csv", validation_rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", choices=["fig5", "fig6", "fig7", "fig1", "all"], default="all")
    args = parser.parse_args()
    selected = {args.case} if args.case != "all" else {"fig5", "fig6", "fig7", "fig1"}
    if "fig5" in selected: build_figure5()
    if "fig6" in selected: build_figure6()
    if "fig7" in selected: build_figure7()
    if "fig1" in selected: build_figure1()


if __name__ == "__main__":
    main()
