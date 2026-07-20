"""Build the Nature Medicine Fig. 4c calibrated heatmap evidence bundle."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
import sys
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import ListedColormap, Normalize
from openpyxl import load_workbook
from PIL import Image, ImageDraw


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from candidate_digitize_heatmap import extract_heatmap


ARTICLE_URL = "https://www.nature.com/articles/s41591-026-04303-y"
FIGURE_URL = "https://www.nature.com/articles/s41591-026-04303-y/figures/4"
IMAGE_URL = (
    "https://media.springernature.com/full/springer-static/image/"
    "art%3A10.1038%2Fs41591-026-04303-y/MediaObjects/41591_2026_4303_Fig4_HTML.png"
)
SOURCE_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41591-026-04303-y/MediaObjects/41591_2026_4303_MOESM3_ESM.xlsx"
)
ROWS = [
    "Z24", "Z31", "Z25", "Z3", "Z2", "Z7", "Z4", "Z23",
    "Z22", "Z30", "Z28", "Z9", "Z12", "Z13", "Z0", "Z6",
    "Z1", "Z17", "Z29", "Z18", "Z26", "Z16", "Z15", "Z19",
    "Z5", "Z11", "Z10", "Z14", "Z27", "Z8", "Z20", "Z21",
]
COLUMNS = [
    ("Age", "Age"),
    ("Sex", "Sex"),
    ("MMSE", "MMSE"),
    ("Aβ42/Aβ40", "CSF_Ab42/40"),
    ("pTau-217", "CSF_pTau217"),
    ("TauPET", "TauPET_MetaROI"),
    ("ADSignCT", "MRI_CTADSign"),
    ("WholeBrainCT", "MRI_WholeBrainCT"),
    ("VentricleVol", "MRI_VentricleVol"),
    ("WMH", "MRI_WMH"),
    ("UPDRS", "UPDRS"),
    ("SAA", "CSF_SAA"),
    ("GFAP", "CSF_GFAP"),
    ("NFL", "CSF_NFL"),
    ("YKL40", "CSF_YKL40"),
    ("sTREM2", "CSF_sTREM2"),
    ("SYT1", "CSF_SYT1"),
    ("SNAP25", "CSF_SNAP25"),
    ("NPTX2", "CSF_NPTX2"),
    ("PDGFRB", "CSF_PDGFRB"),
    ("S100", "CSF_S100"),
]
GRID_BOUNDS = (82.0, 53.5, 751.0, 715.5)
COLORBAR_BOUNDS = (816, 63, 845, 282)
COLORBAR_TOP = 0.3
COLORBAR_BOTTOM = -0.3


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def panel_image(path: Path) -> Image.Image:
    image = Image.open(path).convert("RGB")
    if image.size == (951, 860):
        return image
    if image.width >= 1951 and image.height >= 1510:
        return image.crop((1000, 650, 1951, 1510))
    raise ValueError(
        f"unsupported source dimensions {image.size}; expected full Fig. 4 or 951x860 panel crop"
    )


def run_extraction(original: Path) -> dict:
    result = extract_heatmap(
        original,
        grid_bounds=GRID_BOUNDS,
        row_labels=ROWS,
        column_labels=[display for display, _ in COLUMNS],
        colorbar_bounds=COLORBAR_BOUNDS,
        colorbar_top_value=COLORBAR_TOP,
        colorbar_bottom_value=COLORBAR_BOTTOM,
        cell_margin=3,
        maximum_palette_distance=15,
    )
    if result["status"] != "candidate" or result["cell_count"] != 672:
        raise RuntimeError("Fig. 4c heatmap extraction did not resolve all 672 cells")
    return result


def extraction_rows(extraction: dict) -> list[dict]:
    rows = []
    for cell in extraction["cells"]:
        rows.append(
            {
                "row_index": cell["row_index"],
                "column_index": cell["column_index"],
                "embedding": cell["row_label"],
                "biomarker": cell["column_label"],
                "digitized_correlation": cell["value"],
                "value_status": cell["value_status"],
                "value_interval": json.dumps(cell["value_interval"]),
                "significant_visible": cell["significant_visible"],
                "fill_rgb": ",".join(str(item) for item in cell["fill_rgb"]),
                "palette_pixel_y": cell["palette_pixel_y"],
                "palette_distance_rgb": cell["palette_distance_rgb"],
                "cell_left_pixel": cell["cell_bounds_pixel"][0],
                "cell_top_pixel": cell["cell_bounds_pixel"][1],
                "cell_right_pixel": cell["cell_bounds_pixel"][2],
                "cell_bottom_pixel": cell["cell_bounds_pixel"][3],
                "status": "candidate_visible_geometry",
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def source_table(path: Path) -> tuple[dict, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["SuppData 8"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    table = {
        row[0]: dict(zip(headers, row))
        for row in sheet.iter_rows(min_row=2, values_only=True)
    }
    return table, headers


def validate_source(extraction: dict, source_path: Path) -> tuple[list[dict], dict]:
    source, headers = source_table(source_path)
    mapping = {display: source_stem for display, source_stem in COLUMNS}
    rows = []
    numeric_errors = []
    tp = fp = fn = tn = 0
    censored = 0
    for cell in extraction["cells"]:
        embedding = cell["row_label"]
        display = cell["column_label"]
        stem = mapping[display]
        source_correlation = float(source[embedding][f"{stem}_Corr"])
        source_ci = source[embedding][f"{stem}_CI"]
        source_significant = bool(source[embedding][f"{stem}_Sig"])
        visible_significant = bool(cell["significant_visible"])
        if visible_significant and source_significant:
            tp += 1
        elif visible_significant and not source_significant:
            fp += 1
        elif not visible_significant and source_significant:
            fn += 1
        else:
            tn += 1

        if cell["value_status"] == "clipped_high":
            interval_consistent = source_correlation >= COLORBAR_TOP
            numeric_error = None
            correlation_status = (
                "validated_censored" if interval_consistent else "source_outside_digitized_interval"
            )
            censored += 1
        elif cell["value_status"] == "clipped_low":
            interval_consistent = source_correlation <= COLORBAR_BOTTOM
            numeric_error = None
            correlation_status = (
                "validated_censored" if interval_consistent else "source_outside_digitized_interval"
            )
            censored += 1
        else:
            interval_consistent = None
            numeric_error = float(cell["value"] - source_correlation)
            numeric_errors.append(numeric_error)
            correlation_status = "validated_numeric"
        significance_status = (
            "validated_visible_flag"
            if visible_significant == source_significant
            else "figure_source_flag_mismatch"
        )
        rows.append(
            {
                "embedding": embedding,
                "biomarker_display": display,
                "source_stem": stem,
                "digitized_correlation": cell["value"],
                "digitized_value_status": cell["value_status"],
                "source_correlation_raw": source_correlation,
                "source_ci_raw": source_ci,
                "correlation_error": numeric_error,
                "interval_consistent": interval_consistent,
                "correlation_validation_status": correlation_status,
                "significant_visible": visible_significant,
                "source_significant_raw": source_significant,
                "significance_validation_status": significance_status,
                "source_file": source_path.name,
                "source_sheet": "SuppData 8",
                "source_columns": f"{stem}_Corr|{stem}_CI|{stem}_Sig",
                "mapping_evidence": "Fig. 4c visible row/column labels cross-checked to SuppData 8 headers",
            }
        )

    absolute = np.abs(numeric_errors)
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
    summary = {
        "status": "partial_validated" if fn or fp else "validated",
        "source_rows": 32,
        "visible_cells": 672,
        "matched_cells": len(rows),
        "numeric_cells": len(numeric_errors),
        "endpoint_censored_cells": censored,
        "numeric_mae": float(absolute.mean()),
        "numeric_rmse": float(np.sqrt(np.mean(np.square(numeric_errors)))),
        "numeric_median_absolute_error": float(np.median(absolute)),
        "numeric_p95_absolute_error": float(np.quantile(absolute, 0.95)),
        "numeric_max_absolute_error": float(absolute.max()),
        "significance": {
            "true_positive": tp,
            "false_positive": fp,
            "false_negative": fn,
            "true_negative": tn,
            "precision": precision,
            "recall": recall,
            "f1": f1,
        },
        "mismatch_cells": [
            {
                "embedding": row["embedding"],
                "biomarker": row["biomarker_display"],
                "digitized_visible": row["significant_visible"],
                "source_raw": row["source_significant_raw"],
            }
            for row in rows
            if row["significance_validation_status"] == "figure_source_flag_mismatch"
        ],
        "source_headers": headers,
    }
    return rows, summary


def draw_overlay(original: Path, output: Path, extraction: dict) -> None:
    base = Image.open(original).convert("RGBA")
    layer = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    for cell in extraction["cells"]:
        left, top, right, bottom = cell["cell_bounds_pixel"]
        draw.rectangle((left, top, right, bottom), outline=(0, 151, 136, 175), width=1)
        if cell["significant_visible"]:
            x, y = cell["cell_center_pixel"]
            draw.ellipse((x - 3, y - 3, x + 3, y + 3), outline=(226, 35, 94, 230), width=1)
    bar_left, bar_top, bar_right, bar_bottom = COLORBAR_BOUNDS
    draw.rectangle(
        (bar_left - 2, bar_top - 2, bar_right + 2, bar_bottom + 2),
        outline=(0, 151, 136, 230),
        width=2,
    )
    Image.alpha_composite(base, layer).convert("RGB").save(output)


def draw_recreation(path: Path, extraction: dict) -> None:
    matrix = np.zeros((len(ROWS), len(COLUMNS)), dtype=float)
    stars = np.zeros_like(matrix, dtype=bool)
    for cell in extraction["cells"]:
        matrix[cell["row_index"], cell["column_index"]] = cell["value"]
        stars[cell["row_index"], cell["column_index"]] = cell["significant_visible"]
    # Extracted palette rows run from +0.3 (top/red) to -0.3
    # (bottom/blue), while Matplotlib colormaps run from vmin to vmax.
    palette = np.asarray(extraction["colorbar"]["palette_rgb"])[::-1] / 255.0
    cmap = ListedColormap(palette)
    figure, axis = plt.subplots(figsize=(9.51, 8.6), dpi=100, facecolor="white")
    image = axis.imshow(
        matrix,
        cmap=cmap,
        norm=Normalize(vmin=COLORBAR_BOTTOM, vmax=COLORBAR_TOP),
        interpolation="nearest",
        aspect="auto",
    )
    axis.set_title("Correlation (embeddings, biomarkers)", loc="left", fontsize=18, pad=12)
    axis.set_xticks(range(len(COLUMNS)), [display for display, _ in COLUMNS], rotation=90, fontsize=10)
    axis.set_yticks(range(len(ROWS)), ROWS, fontsize=10)
    axis.set_xticks(np.arange(-0.5, len(COLUMNS), 1), minor=True)
    axis.set_yticks(np.arange(-0.5, len(ROWS), 1), minor=True)
    axis.grid(which="minor", color="#464646", linewidth=0.6)
    axis.tick_params(which="minor", bottom=False, left=False)
    for row in range(len(ROWS)):
        for column in range(len(COLUMNS)):
            if stars[row, column]:
                axis.text(column, row, "*", ha="center", va="center", color="white", fontsize=11, fontweight="bold")
    colorbar = figure.colorbar(image, ax=axis, fraction=0.042, pad=0.08, ticks=[-0.3, 0, 0.3])
    colorbar.set_label("Correlation", rotation=270, labelpad=18, fontsize=12)
    colorbar.outline.set_visible(False)
    figure.subplots_adjust(left=0.10, right=0.88, bottom=0.19, top=0.93)
    figure.savefig(path, facecolor="white")
    plt.close(figure)


def draw_validation(path: Path, rows: list[dict], summary: dict) -> None:
    numeric = [row for row in rows if row["correlation_error"] is not None]
    source = np.asarray([row["source_correlation_raw"] for row in numeric])
    digitized = np.asarray([row["digitized_correlation"] for row in numeric])
    figure, axis = plt.subplots(figsize=(6.2, 6.2), dpi=120, facecolor="white")
    axis.scatter(source, digitized, s=14, color="#1f77b4", alpha=0.65, edgecolors="none")
    axis.plot([-0.31, 0.31], [-0.31, 0.31], color="#222222", linewidth=1)
    axis.set_xlim(-0.31, 0.31)
    axis.set_ylim(-0.31, 0.31)
    axis.set_aspect("equal")
    axis.set_xlabel("Official SuppData 8 correlation")
    axis.set_ylabel("Digitized colour-bar value")
    axis.set_title(f"Numeric cells · MAE {summary['numeric_mae']:.4f}", loc="left")
    axis.spines[["top", "right"]].set_visible(False)
    axis.grid(color="#e7e7e7", linewidth=0.7)
    figure.tight_layout()
    figure.savefig(path, facecolor="white")
    plt.close(figure)


def build_report(
    original: Path,
    source_copy: Path,
    extraction: dict,
    validation: dict,
) -> dict:
    return {
        "schema_version": 1,
        "case_id": "nature-protaide-heatmap",
        "status": "candidate_source_validated",
        "family": "calibrated_heatmap",
        "article": {
            "title": "A deep joint-learning proteomics model for diagnosis of six conditions associated with dementia",
            "journal": "Nature Medicine",
            "year": 2026,
            "doi": "10.1038/s41591-026-04303-y",
            "figure": "Fig. 4c",
            "article_url": ARTICLE_URL,
            "figure_url": FIGURE_URL,
            "image_url": IMAGE_URL,
            "license": "CC BY 4.0",
            "license_url": "https://creativecommons.org/licenses/by/4.0/",
        },
        "input": {
            "file": original.name,
            "sha256": sha256(original),
            "dimensions": list(Image.open(original).size),
            "crop_from_full_figure": [1000, 650, 1951, 1510],
        },
        "recoverable_representation": "32×21 visible cell grid, colour-bar encoded correlation intervals/values, and visible white significance marks",
        "non_recoverable": "exact correlation magnitudes beyond the displayed ±0.3 colour-bar endpoints",
        "candidate_extraction": extraction,
        "source_validation": {
            **validation,
            "source_url": SOURCE_URL,
            "source_file": source_copy.name,
            "source_sha256": sha256(source_copy),
            "source_sheet": "SuppData 8",
            "mapping": [
                {
                    "display_label": display,
                    "source_columns": [f"{stem}_Corr", f"{stem}_CI", f"{stem}_Sig"],
                }
                for display, stem in COLUMNS
            ],
        },
        "webplotdigitizer_comparison": "not_compared",
        "outputs": {
            "digitized_csv": "data.csv",
            "source_validation_csv": "source-validation.csv",
            "source_validation_plot": "source-validation.png",
            "overlay": "overlay.png",
            "recreation": "recreated.png",
        },
        "limitations": [
            "Endpoint-coloured cells are reported as ≥0.3 or ≤−0.3 intervals; source values are not copied into the immutable digitization CSV.",
            "The visible figure omits the white star at Z17–MMSE even though SuppData 8 marks it significant; the mismatch is preserved rather than corrected in the recreation.",
            "The route remains candidate pending held-out heatmap styles and a fair same-input WebPlotDigitizer comparison.",
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--source-data", type=Path, required=True)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=ROOT / "gallery" / "assets" / "cases" / "nature-protaide-heatmap",
    )
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    original = args.output_dir / "original.png"
    source_copy = args.output_dir / "source-data.xlsx"
    panel_image(args.input).save(original)
    if args.source_data.resolve() != source_copy.resolve():
        shutil.copy2(args.source_data, source_copy)
    extraction = run_extraction(original)
    rows = extraction_rows(extraction)
    write_csv(args.output_dir / "data.csv", rows)
    validation_rows, validation_summary = validate_source(extraction, source_copy)
    write_csv(args.output_dir / "source-validation.csv", validation_rows)
    draw_overlay(original, args.output_dir / "overlay.png", extraction)
    draw_recreation(args.output_dir / "recreated.png", extraction)
    draw_validation(args.output_dir / "source-validation.png", validation_rows, validation_summary)
    report = build_report(original, source_copy, extraction, validation_summary)
    (args.output_dir / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"CASE={args.output_dir}")
    print("STATUS=candidate_source_validated")
    print(f"CELLS={extraction['cell_count']}/672")
    print(f"NUMERIC_MAE={validation_summary['numeric_mae']:.6f}")


if __name__ == "__main__":
    main()
