"""Build the Nature Communications Fig. 4d dose-response gallery case."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import shutil
import sys
from pathlib import Path

import fitz
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from PIL import Image, ImageDraw


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from candidate_digitize_dose_response_pdf import (  # noqa: E402
    LinearCalibration,
    SeriesSpec,
    extract_dose_response_pdf,
)
from inspect_pdf_vectors import inspect_page  # noqa: E402


CASE_ID = "nature-kahlous-dose-response"
ARTICLE_URL = "https://www.nature.com/articles/s41467-026-71361-8"
FIGURE_URL = f"{ARTICLE_URL}/figures/4"
IMAGE_URL = (
    "https://media.springernature.com/full/springer-static/image/"
    "art%3A10.1038%2Fs41467-026-71361-8/MediaObjects/"
    "41467_2026_71361_Fig4_HTML.png"
)
PDF_URL = f"{ARTICLE_URL}.pdf"
SOURCE_DATA_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41467-026-71361-8/MediaObjects/41467_2026_71361_MOESM7_ESM.xlsx"
)

PDF_PAGE = 7
PANEL_ROI = (55.0, 178.0, 235.0, 310.0)
MAIN_PLOT_ROI = (106.7, 213.8, 192.3, 280.8)
RENDER_SCALE = 4.0
X_CALIBRATION = LinearCalibration(116.836998, -10.0, 192.264999, -2.0)
Y_CALIBRATION = LinearCalibration(277.919006, 0.0, 222.200012, 100.0)
SERIES_SPECS = [
    SeriesSpec("ADR", "square", "#ffba12", "#ffba10"),
    SeriesSpec("NA", "triangle", "#748d8e", "#758c8e"),
    SeriesSpec("DA", "circle", "#48e99d", "#48ea9c"),
]
PLOT_COLOURS = {"ADR": "#f2a900", "NA": "#748d8e", "DA": "#32d98b"}
MARKERS = {"ADR": "s", "NA": "^", "DA": "o"}
SOURCE_COLUMNS = {"ADR": (2, 3, 4), "NA": (5, 6, 7), "DA": (8, 9, 10)}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def source_table(path: Path) -> dict[tuple[str, int], dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Figure 4"]
    anchor_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == "D5R_Human_M74":
            anchor_row = row
            break
    if anchor_row is None:
        raise RuntimeError("D5R_Human_M74 block not found in Figure 4 source-data sheet")

    table: dict[tuple[str, int], dict] = {}
    for row in range(anchor_row + 3, anchor_row + 11):
        log10_molar = sheet.cell(row, 1).value
        if not isinstance(log10_molar, (int, float)):
            continue
        for series, (mean_col, sd_col, n_col) in SOURCE_COLUMNS.items():
            mean = sheet.cell(row, mean_col).value
            sd = sheet.cell(row, sd_col).value
            n = sheet.cell(row, n_col).value
            if mean is None or sd is None or n is None:
                continue
            table[(series, int(log10_molar))] = {
                "source_mean": float(mean) * 100,
                "source_sd": float(sd) * 100,
                "source_n": int(n),
                "source_sem": float(sd) / math.sqrt(float(n)) * 100,
                "source_sheet": "Figure 4",
                "source_block": "D5R_Human_M74",
                "source_row": row,
            }
    if len(table) != 18:
        raise RuntimeError(f"expected 18 populated source means, found {len(table)}")
    return table


def validate_source(extraction: dict, source_path: Path) -> tuple[list[dict], dict]:
    source = source_table(source_path)
    rows = []
    marker_errors = []
    error_endpoint_errors = []
    for point in extraction["points"]:
        if point["segment"] == "vehicle":
            rows.append(
                {
                    "series": point["series"],
                    "segment": "vehicle",
                    "digitized_log10_molar": "",
                    "source_log10_molar": "",
                    "digitized_value": point["plotted_value"],
                    "source_mean": "",
                    "marker_error": "",
                    "digitized_error_lower": "",
                    "digitized_error_upper": "",
                    "source_error_lower": "",
                    "source_error_upper": "",
                    "source_sem": "",
                    "source_n": "",
                    "marker_status": "vector_marker_extracted",
                    "error_status": "not_visible",
                    "source_status": "vehicle_not_present_in_source_workbook",
                    "source_sheet": "Figure 4",
                    "source_row": "",
                }
            )
            continue

        source_log = int(round(point["log10_molar"]))
        source_row = source[(point["series"], source_log)]
        source_mean = source_row["source_mean"]
        source_sem = source_row["source_sem"]
        marker_error = float(point["plotted_value"] - source_mean)
        marker_errors.append(marker_error)
        error_visible = point["error_lower"] is not None
        lower_source = source_mean - source_sem
        upper_source = source_mean + source_sem
        if error_visible:
            lower_error = float(point["error_lower"] - lower_source)
            upper_error = float(point["error_upper"] - upper_source)
            error_endpoint_errors.extend([lower_error, upper_error])
            error_status = "visible_sem_endpoints_validated"
        else:
            lower_error = upper_error = None
            error_status = "sem_occluded_by_marker"
        rows.append(
            {
                "series": point["series"],
                "segment": "main",
                "digitized_log10_molar": point["log10_molar"],
                "source_log10_molar": source_log,
                "digitized_value": point["plotted_value"],
                "source_mean": source_mean,
                "marker_error": marker_error,
                "digitized_error_lower": point["error_lower"] if error_visible else "",
                "digitized_error_upper": point["error_upper"] if error_visible else "",
                "source_error_lower": lower_source,
                "source_error_upper": upper_source,
                "source_sem": source_sem,
                "source_n": source_row["source_n"],
                "marker_status": "vector_marker_extracted",
                "error_status": error_status,
                "source_status": "official_source_mean_sem_mapped",
                "source_sheet": source_row["source_sheet"],
                "source_row": source_row["source_row"],
                "error_lower_residual": lower_error if lower_error is not None else "",
                "error_upper_residual": upper_error if upper_error is not None else "",
            }
        )

    marker_absolute = np.abs(np.asarray(marker_errors, dtype=float))
    endpoint_absolute = np.abs(np.asarray(error_endpoint_errors, dtype=float))
    summary = {
        "status": "validated_against_official_source_data",
        "visible_markers": len(extraction["points"]),
        "source_matched_markers": len(marker_errors),
        "source_uncovered_vehicle_markers": 3,
        "visible_error_bars": int(extraction["summary"]["visible_error_bar_count"]),
        "source_sem_values": 18,
        "sem_occluded_by_marker": 18 - int(extraction["summary"]["visible_error_bar_count"]),
        "marker_mae": float(marker_absolute.mean()),
        "marker_max_absolute_error": float(marker_absolute.max()),
        "visible_error_endpoint_mae": float(endpoint_absolute.mean()),
        "visible_error_endpoint_max_absolute_error": float(endpoint_absolute.max()),
        "source_value_transform": "workbook mean and SD multiplied by 100; SEM = SD / sqrt(N)",
        "source_role": "independent validation; never substituted for digitized PDF geometry",
    }
    return rows, summary


def write_csv(path: Path, rows: list[dict]) -> None:
    fields = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def extraction_rows(extraction: dict, validation_rows: list[dict]) -> list[dict]:
    validation = {
        (row["series"], row["segment"], row["source_log10_molar"]): row
        for row in validation_rows
    }
    rows = []
    for point in extraction["points"]:
        source_log = "" if point["segment"] == "vehicle" else int(round(point["log10_molar"]))
        checked = validation[(point["series"], point["segment"], source_log)]
        rows.append(
            {
                "series": point["series"],
                "marker_shape": point["marker_shape"],
                "segment": point["segment"],
                "log10_molar": "" if point["log10_molar"] is None else point["log10_molar"],
                "digitized_value": point["plotted_value"],
                "digitized_error_lower": "" if point["error_lower"] is None else point["error_lower"],
                "digitized_error_upper": "" if point["error_upper"] is None else point["error_upper"],
                "pdf_x_pt": point["pdf_x_pt"],
                "pdf_y_pt": point["pdf_y_pt"],
                "status": point["status"],
                "source_mean": checked["source_mean"],
                "source_sem": checked["source_sem"],
                "marker_error": checked["marker_error"],
                "source_status": checked["source_status"],
                "error_status": checked["error_status"],
            }
        )
    return rows


def write_curves_csv(path: Path, extraction: dict) -> None:
    rows = []
    for curve in extraction["curves"]:
        for sample_index, point in enumerate(curve["points"]):
            rows.append(
                {
                    "series": curve["series"],
                    "sample_index": sample_index,
                    "log10_molar": point["log10_molar"],
                    "plotted_value": point["plotted_value"],
                    "pdf_x_pt": point["pdf_x_pt"],
                    "pdf_y_pt": point["pdf_y_pt"],
                    "status": "curve_path_traced",
                }
            )
    write_csv(path, rows)


def render_original_and_overlay(pdf_path: Path, extraction: dict, output_dir: Path) -> None:
    with fitz.open(pdf_path) as document:
        pixmap = document[PDF_PAGE - 1].get_pixmap(
            matrix=fitz.Matrix(RENDER_SCALE, RENDER_SCALE),
            clip=fitz.Rect(*PANEL_ROI),
            alpha=False,
        )
        pixmap.save(output_dir / "original.png")

    original = Image.open(output_dir / "original.png").convert("RGBA")
    overlay = Image.new("RGBA", original.size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(overlay)

    def pixel(pdf_x: float, pdf_y: float) -> tuple[float, float]:
        return (
            (pdf_x - PANEL_ROI[0]) * RENDER_SCALE,
            (pdf_y - PANEL_ROI[1]) * RENDER_SCALE,
        )

    for curve in extraction["curves"]:
        coordinates = [pixel(point["pdf_x_pt"], point["pdf_y_pt"]) for point in curve["points"]]
        if len(coordinates) > 1:
            draw.line(coordinates, fill=(0, 117, 170, 135), width=2)

    for point in extraction["points"]:
        x, y = pixel(point["pdf_x_pt"], point["pdf_y_pt"])
        radius = 11
        draw.ellipse((x - radius, y - radius, x + radius, y + radius), outline=(0, 117, 170, 255), width=3)
        if point["error_lower"] is not None:
            for error_value in (point["error_lower"], point["error_upper"]):
                error_y_pt = Y_CALIBRATION.pixel_1 + (
                    error_value - Y_CALIBRATION.value_1
                ) * (Y_CALIBRATION.pixel_2 - Y_CALIBRATION.pixel_1) / (
                    Y_CALIBRATION.value_2 - Y_CALIBRATION.value_1
                )
                _, error_y = pixel(point["pdf_x_pt"], error_y_pt)
                draw.line((x - 7, error_y, x + 7, error_y), fill=(209, 54, 103, 255), width=2)

    composed = Image.alpha_composite(original, overlay).convert("RGB")
    composed.save(output_dir / "overlay.png")


def recreate(extraction: dict, output_path: Path) -> None:
    figure = plt.figure(figsize=(6.8, 4.8), dpi=140, facecolor="white")
    grid = figure.add_gridspec(1, 2, width_ratios=[0.8, 5.0], wspace=0.08)
    vehicle_axis = figure.add_subplot(grid[0, 0])
    main_axis = figure.add_subplot(grid[0, 1], sharey=vehicle_axis)

    for axis in (vehicle_axis, main_axis):
        axis.set_facecolor("white")
        axis.spines[["top", "right"]].set_visible(False)
        axis.tick_params(direction="out", width=1.0, length=4)
    vehicle_axis.spines["right"].set_visible(False)
    main_axis.spines["left"].set_visible(False)
    main_axis.tick_params(axis="y", left=False, labelleft=False)
    vehicle_axis.set_xlim(-0.08, 0.35)
    vehicle_axis.set_xticks([0])
    vehicle_axis.set_xticklabels(["0"])
    main_axis.set_xlim(-11.1, -2.0)
    main_axis.set_xticks([-10, -8, -6, -4, -2])
    main_axis.set_xticks([-11, -9, -7, -5, -3], minor=True)
    vehicle_axis.set_ylim(0, 110)
    vehicle_axis.set_yticks([0, 50, 100])
    vehicle_axis.set_ylabel("% of max. cAMP")
    main_axis.set_xlabel("log [ligand], M")
    main_axis.set_title("D$_5$R · M74", loc="left", fontsize=12, pad=10)

    curves = {curve["series"]: curve for curve in extraction["curves"]}
    for spec in SERIES_SPECS:
        colour = PLOT_COLOURS[spec.name]
        points = [point for point in extraction["points"] if point["series"] == spec.name]
        vehicle = next(point for point in points if point["segment"] == "vehicle")
        main = sorted(
            (point for point in points if point["segment"] == "main"),
            key=lambda point: point["log10_molar"],
        )
        curve = curves[spec.name]["points"]
        main_axis.plot(
            [point["log10_molar"] for point in curve],
            [point["plotted_value"] for point in curve],
            color=colour,
            linewidth=1.6,
        )
        for point in main:
            lower = point["error_lower"]
            upper = point["error_upper"]
            if lower is not None:
                main_axis.errorbar(
                    [point["log10_molar"]],
                    [point["plotted_value"]],
                    yerr=[[point["plotted_value"] - lower], [upper - point["plotted_value"]]],
                    fmt="none",
                    ecolor=colour,
                    elinewidth=1.0,
                    capsize=2.2,
                )
        main_axis.plot(
            [point["log10_molar"] for point in main],
            [point["plotted_value"] for point in main],
            linestyle="none",
            marker=MARKERS[spec.name],
            markersize=5.8,
            markerfacecolor="white",
            markeredgecolor=colour,
            markeredgewidth=1.2,
        )
        vehicle_axis.plot([0, 0.3], [vehicle["plotted_value"]] * 2, color=colour, linewidth=1.4)
        vehicle_axis.plot(
            [0],
            [vehicle["plotted_value"]],
            linestyle="none",
            marker=MARKERS[spec.name],
            markersize=5.8,
            markerfacecolor="white",
            markeredgecolor=colour,
            markeredgewidth=1.2,
        )

    figure.savefig(output_path, facecolor="white", bbox_inches="tight")
    plt.close(figure)


def validation_plot(rows: list[dict], output_path: Path) -> None:
    matched = [row for row in rows if row["source_status"] == "official_source_mean_sem_mapped"]
    figure, axis = plt.subplots(figsize=(5.2, 4.2), dpi=140, facecolor="white")
    for series in [spec.name for spec in SERIES_SPECS]:
        series_rows = [row for row in matched if row["series"] == series]
        axis.scatter(
            [row["source_mean"] for row in series_rows],
            [row["digitized_value"] for row in series_rows],
            color=PLOT_COLOURS[series],
            label=series,
            s=32,
        )
    axis.plot([15, 95], [15, 95], color="#222222", linewidth=1, linestyle="--")
    axis.set_xlim(15, 95)
    axis.set_ylim(15, 95)
    axis.set_xlabel("Official source mean (%)")
    axis.set_ylabel("Digitized vector marker (%)")
    axis.spines[["top", "right"]].set_visible(False)
    axis.legend(frameon=False, ncols=3, fontsize=8)
    figure.savefig(output_path, facecolor="white", bbox_inches="tight")
    plt.close(figure)


def copy_if_needed(source: Path, target: Path) -> None:
    if source.resolve() != target.resolve():
        shutil.copy2(source, target)


def build(pdf_path: Path, source_path: Path, output_dir: Path) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    extraction = extract_dose_response_pdf(
        pdf_path,
        page_number=PDF_PAGE,
        panel_roi=PANEL_ROI,
        main_plot_roi=MAIN_PLOT_ROI,
        x_calibration=X_CALIBRATION,
        y_calibration=Y_CALIBRATION,
        series_specs=SERIES_SPECS,
    )
    expected = {
        "visible_marker_count": 21,
        "main_marker_count": 18,
        "vehicle_marker_count": 3,
        "visible_error_bar_count": 14,
        "traced_curve_count": 3,
    }
    if extraction["summary"] != expected:
        raise RuntimeError(f"unexpected Fig. 4d extraction summary: {extraction['summary']}")

    validation_rows, validation = validate_source(extraction, source_path)
    if validation["marker_max_absolute_error"] >= 0.1:
        raise RuntimeError("Fig. 4d marker validation exceeded 0.1 percentage points")
    if validation["visible_error_endpoint_max_absolute_error"] >= 0.2:
        raise RuntimeError("Fig. 4d error-endpoint validation exceeded 0.2 percentage points")

    render_original_and_overlay(pdf_path, extraction, output_dir)
    recreate(extraction, output_dir / "recreated.png")
    write_csv(output_dir / "data.csv", extraction_rows(extraction, validation_rows))
    write_curves_csv(output_dir / "curves.csv", extraction)
    write_csv(output_dir / "source-validation.csv", validation_rows)
    validation_plot(validation_rows, output_dir / "source-validation.png")
    copy_if_needed(pdf_path, output_dir / "source-article.pdf")
    copy_if_needed(source_path, output_dir / "source-data.xlsx")

    with fitz.open(pdf_path) as document:
        inspection = {
            "schema_version": 1,
            "input_sha256": sha256(pdf_path),
            **inspect_page(document, PDF_PAGE - 1),
        }
    (output_dir / "vector-inspection.json").write_text(
        json.dumps(inspection, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    report = {
        "schema_version": 1,
        "case_id": CASE_ID,
        "status": "candidate",
        "family": "dose_response",
        "title": "D5R M74 catecholamine concentration-response curves",
        "article": {
            "title": "Molecular mechanisms of native ligand selectivity in catecholamine G protein-coupled receptors",
            "journal": "Nature Communications",
            "year": 2026,
            "doi": "10.1038/s41467-026-71361-8",
            "article_url": ARTICLE_URL,
            "figure_url": FIGURE_URL,
            "image_url": IMAGE_URL,
            "pdf_url": PDF_URL,
            "source_data_url": SOURCE_DATA_URL,
            "figure": "Fig. 4d",
            "license": "CC BY 4.0",
        },
        "inputs": {
            "pdf_sha256": sha256(pdf_path),
            "source_data_sha256": sha256(source_path),
            "pdf_page": PDF_PAGE,
            "panel_roi_pt": list(PANEL_ROI),
        },
        "candidate_extraction": extraction,
        "source_validation": validation,
        "claims": {
            "points": "vector_marker_extracted",
            "visible_error_bars": "vector_error_endpoints_extracted",
            "curves": "curve_path_traced",
            "curve_parameters": "not_extracted",
            "vehicle_source_values": "not_present_in_source_workbook",
            "raw_replicates": "not_recoverable_from_rendered_panel",
        },
        "limitations": extraction["limitations"]
        + [
            "Four source-data SEM intervals are fully occluded by their visible markers and are not claimed as image-extracted error bars.",
            "The official workbook validates the 18 concentration points but contains no vehicle rows for this panel.",
            "This is a verified real-figure candidate case, not yet a stable general dose-response extractor.",
            "WebPlotDigitizer has not yet been compared under matched calibration and intervention conditions.",
        ],
    }
    (output_dir / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pdf", required=True, type=Path)
    parser.add_argument("--source-data", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()
    report = build(args.pdf, args.source_data, args.output_dir)
    print(json.dumps(report["source_validation"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
